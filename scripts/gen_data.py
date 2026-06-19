#!/usr/bin/env python3
"""Gera assets/data/iniciativas.js a partir de PTE2026_matriz_dashboard.xlsx.
Uso: python3 scripts/gen_data.py
"""
import openpyxl, json, os

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "PTE2026_matriz_dashboard.xlsx")
OUT = os.path.join(ROOT, "assets", "data", "iniciativas.js")

# Pré-selecionadas (aba Cruzamento): id -> UFs onde aparece
PRE = {27: ["MA"], 21: ["PI"], 52: ["PI", "CE"], 10: ["CE", "RN", "PB", "PE", "AL", "BA"],
       20: ["CE"], 74: ["CE"], 32: ["CE"], 38: ["CE"], 57: ["CE"], 58: ["CE"], 6: ["RN"],
       34: ["RN"], 24: ["PB", "PE"], 76: ["PB"], 53: ["PB"], 1: ["PE"], 12: ["PE"], 13: ["PE"],
       39: ["PE"], 56: ["AL"], 9: ["AL"], 16: ["AL"], 66: ["AL"], 67: ["AL"], 47: ["AL"],
       50: ["SE"], 75: ["SE"], 3: ["BA"], 17: ["BA"], 65: ["BA"], 72: ["BA"]}
FORA_NE = {11, 19, 59, 60, 70}  # sede fora do NE (não são sites de visita de campo)
EXCLUDE = {29, 8}  # 29: mesma entidade da ID 56 · 8: Projeto Pacto Global de Jovens pelo Clima (a pedido)

# Iniciativas com mais de um local de visita possível. A rota escolhe a
# coordenada ótima entre estes pontos; o mapa mostra todos.
ALT_LOCAIS = {
    24: [  # Programa 1 Milhão de Tetos Solares
        {"municipio": "Remígio", "uf": "PB", "lat": -6.9528, "lon": -35.8331},
        {"municipio": "Araripina", "uf": "PE", "lat": -7.5764, "lon": -40.4983},
    ],
}

# Revisão editorial dos nomes: nome principal padronizado, sem siglas de UF,
# sem subtítulos/descrições e sem acrônimo da organização (que vira subtítulo).
NAME_OVERRIDES = {
    3: "Conselho Gestor do Fundo Rotativo",
    5: "Rede de Ativadores de Crédito Socioambiental",
    7: "ID Hub Brazil",
    8: "Projeto Pacto Global de Jovens pelo Clima",
    9: "OxeTech",
    11: "Programa Jovem Empreendedor Primeiros Passos (JEPP)",
    13: "Parque Tecnológico Porto Digital",
    17: "Rede BATUC de Turismo Comunitário",
    18: "Associação das Comunidades Negras Rurais Quilombolas do Maranhão",
    19: "Territórios da Cidadania",
    21: "Green Energy Park",
    22: "Logística Verde e Operações Sustentáveis no Porto de Suape",
    23: "SENAI CIMATEC",
    24: "Programa 1 Milhão de Tetos Solares (P1MTS)",
    25: "Programa Água Doce",
    26: "APROBAMBU",
    27: "Raízes Solares",
    28: "Programa Sertão Vivo",
    29: "Crédito de Carbono Integral (CCI-BSHE)",
    30: "Programa Município Selo Verde",
    34: "Modelo de Gestão Municipal de Resíduos Sólidos de Arez",
    37: "Labifor",
    39: "Programas Complementares de Captação de Água da ASA",
    40: "Projeto Tecnologia SARA",
    41: "Zoneamento Ecológico-Econômico da Zona Costeira do Ceará",
    44: "Conecta Caatinga",
    47: "SIMACaatinga",
    49: "Renova-Semiárido",
    51: "No Clima da Caatinga",
    56: "Projetos de Conservação e Sustentabilidade do Bioma Caatinga",
    59: "Comitê da Bacia Hidrográfica do Rio São Francisco (CBHSF)",
    60: "Observatório da Transição Energética",
    61: "Hub de Hidrogênio Verde do Complexo do Pecém",
    62: "Projeto Pecém – Fortescue",
    68: "Voltalia – Cluster Serra Branca",
    69: "Complexo Solar São Gonçalo – Enel Green Power",
    72: "Programa Indústria Verde – FIEB",
    73: "Plano de Descarbonização e Hub de H2V – Porto do Itaqui / EMAP",
    74: "Complexo Eólico Marinho Dragão do Mar – Energo",
    75: "Hub de Hidrogênio Verde de Sergipe – Green Energy Park / SergipeTec",
    76: "Cooperativa de Energia Solar Bem Viver / CERSA",
    77: "Grupo EQM / ZEG Biogás – Biometano da Vinhaça",
    78: "SENAI e Hubs de Inovação em Hidrogênio Verde / Powershoring",
    80: "Debêntures Verdes da Casa dos Ventos (Complexo Rio do Vento)",
    81: "Instituto Clima e Sociedade",
    82: "SITAWI Finanças do Bem",
}

# Ajuste de organização (subtítulo) — quando difere do que consta na planilha
ORG_OVERRIDES = {
    56: "Associação de Produtores de Crédito da Caatinga",
}

# Resumo editorial (1–2 frases) exibido nos cartões; o objetivo completo é preservado nos dados
RESUMO = {
    1: "Oferece microcrédito produtivo orientado, educação financeira e formalização (MEI) a empreendedores do Agreste e da Zona da Mata de Pernambuco, com forte protagonismo de mulheres e jovens.",
    2: "Banco comunitário do Conjunto Palmeiras que democratiza serviços financeiros à população de baixa renda por meio de microcrédito, moeda social e apoio a empreendedores locais.",
    3: "Articula e gere fundos rotativos solidários em rede, fortalecendo cooperativas de crédito, associações comunitárias e a agricultura familiar com capital de giro.",
    4: "Promove inclusão financeira na Região Metropolitana do Recife pela moeda social Caiana, estimulando o comércio local e o acesso a crédito popular.",
    5: "Forma uma rede de técnicos locais para destravar o acesso ao crédito rural (Pronaf) de agricultores familiares e comunidades tradicionais, unindo assistência técnica e educação financeira.",
    6: "Organiza agricultores familiares e empreendimentos solidários para produzir e comercializar alimentos agroecológicos com certificação orgânica participativa, articulando agroecologia, feminismo e juventude.",
    7: "Centro de Descarbonização da Indústria, fruto de cooperação Brasil–Reino Unido, que apoia projetos industriais na transição para uma economia de baixo carbono.",
    8: "Prepara jovens estudantes, pela educação ambiental ativa, para atuar como protagonistas na luta por justiça socioambiental diante das crises climáticas.",
    9: "Programa alagoano de capacitação técnica em tecnologia que fomenta o empreendedorismo local e a atração de investimentos.",
    10: "Fomenta a competitividade da indústria brasileira pela adoção de inteligência artificial e tecnologias de fronteira nos processos produtivos.",
    11: "Estimula criatividade, pensamento crítico e mentalidade empreendedora em alunos do ensino fundamental, de forma lúdica e com temas como a sustentabilidade.",
    12: "Atua na conservação e recuperação da Caatinga em territórios tradicionais, com tecnologias sociais e inclusão produtiva das comunidades.",
    13: "Ecossistema de inovação que integra empresas, governo e universidades, dinamizando projetos em TI e economia criativa.",
    14: "Fomenta ecossistemas de inovação com foco em diversidade, atuando na aceleração de negócios, em eventos e em hubs de apoio a empreendedores.",
    15: "Fortalece a agricultura familiar no Semiárido, ampliando resiliência climática, segurança hídrica e inclusão produtiva com práticas agroecológicas e infraestrutura rural.",
    16: "Valoriza saberes tradicionais e produtos da sociobiodiversidade criando rotas de turismo de base comunitária que geram renda e fortalecem identidades locais.",
    17: "Promove o turismo comunitário em comunidades tradicionais da Bahia, fortalecendo economias locais, cultura e conservação ambiental.",
    18: "Fortalece a autonomia produtiva e territorial de mulheres quilombolas com práticas agroecológicas, uso sustentável da sociobiodiversidade e cadeias solidárias.",
    19: "Promove o desenvolvimento territorial sustentável no Nordeste, articulando políticas públicas para inclusão produtiva e acesso a direitos em territórios rurais e vulnerabilizados.",
    20: "Promove uma economia de baixo carbono no Semiárido pelo uso sustentável da Caatinga, restauração ecológica e fortalecimento de cadeias da sociobiodiversidade.",
    21: "Polo de produção e exportação de hidrogênio verde e amônia verde no litoral do Piauí, aproveitando o potencial eólico e solar para descarbonizar indústria e transporte marítimo.",
    22: "Transforma o complexo portuário de Suape em polo de energia verde e corredor logístico descarbonizado, atraindo projetos de hidrogênio verde e derivados.",
    23: "Complexo de educação profissional, pesquisa e inovação em tecnologias limpas — hidrogênio verde, bioenergia e smart grids — para a transição energética do Nordeste.",
    24: "Promove energia fotovoltaica descentralizada em residências e unidades comunitárias do Semiárido, fortalecendo a autonomia energética e a justiça climática para famílias de baixa renda.",
    25: "Dessaliniza águas salobras para garantir abastecimento seguro a comunidades rurais do Semiárido, integrando a solução à geração de energia renovável descentralizada.",
    26: "Desenvolve uma cadeia de bioeconomia do bambu para etanol de segunda geração, biomassa energética e bioprodutos, aliando inovação e desenvolvimento rural.",
    27: "Implanta uma usina solar comunitária em território quilombola para reduzir a pobreza energética e criar um fundo econômico comunitário autossustentável.",
    28: "Integra energia solar ao bombeamento de água e à agroecologia no Semiárido, substituindo o diesel para reduzir custos e emissões e garantir água e energia à produção familiar.",
    30: "Certificação ambiental pública que reconhece municípios com políticas efetivas de conservação e uso sustentável dos recursos naturais.",
    31: "Articula políticas de inclusão produtiva e inovação social para fortalecer empreendimentos econômicos solidários de base comunitária e cooperativa na Bahia.",
    32: "Startup que desenvolve soluções para o descarte correto de resíduos, fortalecendo cooperativas e aprimorando a logística reversa.",
    33: "Amplia a coleta seletiva e a reciclagem de resíduos sólidos e valoriza o trabalho dos catadores como agentes centrais da gestão de resíduos.",
    34: "Desativa lixão e implanta gestão integrada de resíduos com suporte técnico e financeiro, reduzindo impactos ambientais e promovendo soluções circulares.",
    35: "Rede que apoia cooperativas de reciclagem com capacitação, digitalização da gestão e elaboração de planos municipais de resíduos, sobretudo em pequenos municípios.",
    36: "Promove a gestão integrada e sustentável dos resíduos sólidos nos municípios da Ibiapaba por meio da cooperação intermunicipal.",
    37: "Laboratório de inovação da Prefeitura de Fortaleza que cria soluções urbanas sustentáveis, como compostagem por biodigestores, microparques e educação ambiental.",
    38: "Integra desenvolvimento urbano e sustentabilidade ambiental, promovendo a resiliência das cidades e o uso eficiente dos recursos naturais.",
    39: "Garante acesso à água para consumo, produção de alimentos e educação no Semiárido por meio de tecnologias sociais de captação.",
    40: "Promove saneamento rural e segurança hídrica com tratamento de esgoto e reúso agrícola da água.",
    41: "Ordena o uso do território costeiro cearense para conservar o capital natural e promover desenvolvimento sustentável e melhoria de vida da população.",
    42: "Oferece transporte público sustentável e de baixo carbono.",
    43: "Recupera áreas degradadas e promove o desenvolvimento territorial integrado.",
    44: "Cria corredores ecológicos conectando áreas protegidas públicas e privadas da Caatinga.",
    45: "Promove a adoção de tecnologias de agricultura de baixa emissão de carbono (ABC) no Semiárido.",
    46: "Promove o empoderamento político e econômico de mulheres rurais aliado à resiliência climática.",
    47: "Sistema de monitoramento contínuo da cobertura vegetal do bioma Caatinga.",
    48: "Promove a produção de algodão orgânico e colorido no Semiárido com agricultura regenerativa, conectando agricultores familiares ao mercado de moda sustentável.",
    49: "Plataforma digital que mapeia casos de sucesso em energias renováveis e tecnologias sociais no Semiárido, com mapa interativo, informações e depoimentos.",
    50: "Pesquisa e desenvolve bioinsumos para a agropecuária — biofertilizantes, biopesticidas e probióticos — do laboratório ao campo.",
    51: "Contribui para a mitigação e a adaptação climática de comunidades rurais conservando a Caatinga, os recursos hídricos e a biodiversidade, em sete linhas de ação.",
    52: "Promove o turismo de natureza e a valorização do patrimônio cultural, integrando comunidades tradicionais e conservando a biodiversidade.",
    53: "Produz alimentos e fibras de forma regenerativa e biodinâmica, restaurando o solo, conservando a Caatinga e sequestrando carbono no Semiárido paraibano.",
    54: "Busca recuperar 10 milhões de hectares degradados na Caatinga até 2045, com protagonismo das comunidades e conservação do solo, da água e da biodiversidade.",
    55: "Maior programa de infraestrutura e apoio direto ao produtor rural de Pernambuco, com mais de R$ 200 milhões para modernizar a agricultura familiar.",
    56: "Reúne agricultores familiares e comunidades tradicionais em um modelo de crédito de carbono integral que gera renda pela floresta em pé, do cânion do São Francisco a todo o Semiárido.",
    57: "Produz matéria-prima orgânica (sobretudo acerola) para suplementos, com rastreabilidade total e geração de renda em parceria com agricultores locais.",
    58: "Cultiva macroalgas nativas com comunidades marinhas tradicionais, gerando renda e conservando a biodiversidade, com produtos para os setores alimentício, cosmético e farmacêutico.",
    59: "Faz a gestão participativa dos recursos hídricos da bacia do Rio São Francisco, promovendo uso sustentável, proteção dos mananciais e mediação de conflitos.",
    60: "Plataforma que cruza a geolocalização de empreendimentos de energia e mineração com milhares de territórios socioambientais para apontar áreas impactadas, sob a ótica da justiça climática.",
    61: "Hub de hidrogênio e amônia verde na ZPE do Pecém, em parceria com o Porto de Roterdã; reúne pré-contratos de cerca de US$ 24 bilhões e é o projeto de H2V mais avançado do país.",
    62: "Planta de hidrogênio e amônia verde na ZPE Ceará, com 175 mil t/ano de H2V usando 1,2 GW renovável e investimento de cerca de R$ 20 bilhões.",
    63: "Primeira planta de hidrogênio verde em escala industrial do Brasil, no Polo de Camaçari, com até US$ 1,5 bilhão para chegar a 100 mil t/ano de H2V até 2027.",
    64: "Transforma o porto de Suape em polo de hidrogênio verde, e-metanol e combustíveis de baixo carbono, com mais de R$ 15,8 bilhões previstos e um TechHub de P&D e capacitação.",
    65: "Maior biorrefinaria de SAF e diesel verde da América Latina (1 bilhão de litros/ano), na Bahia, com cultivo de macaúba em pastagens degradadas.",
    66: "Complexo de biorrefinarias em Alagoas que produz etanol de baixo carbono, biometano, CO2 biogênico e biofertilizantes a partir de resíduos da cana.",
    67: "Maior produtora de etanol de Alagoas que, com a Veolia, biodigere a vinhaça para gerar biogás, vapor verde e biometano, tornando-se autossuficiente em energia.",
    68: "Maior complexo eólico-solar da Voltalia no mundo, no Rio Grande do Norte, com potencial de 2,4 GW e mais de R$ 3 bilhões investidos, abrindo espaço para hidrogênio verde.",
    69: "Um dos maiores parques solares da América do Sul (864 MW), no Piauí, gerando mais de 2 TWh/ano e evitando cerca de 1,3 milhão de toneladas de CO2.",
    70: "Líder brasileira em geração distribuída solar compartilhada, com mais de 200 fazendas solares e 110 mil clientes pelo modelo de assinatura sem instalação.",
    71: "Plataforma do Consórcio Nordeste, Banco do Nordeste e iCS para coordenar a atração de cadeias produtivas verdes baseadas no potencial energético da região.",
    72: "Programa do Sistema FIEB que apoia cadeias sustentáveis, bioeconomia e transição energética na Bahia, estado que concentra R$ 78 bilhões em investimentos verdes previstos.",
    73: "Plano de descarbonização do Porto do Itaqui rumo à neutralidade até 2050, com projetos de hidrogênio e aço verde aproveitando a logística e o potencial renovável do Maranhão.",
    74: "Projeto pioneiro de energia eólica offshore no Ceará, com R$ 18 bilhões previstos e potencial de 4 GW, em fase de licenciamento ambiental.",
    75: "Hub de hidrogênio verde coordenado pela Desenvolve-SE, com até US$ 5 bilhões previstos, integrado ao terminal marítimo e à ZPE, com pesquisa em solar e biogás.",
    76: "Cooperativa de energia solar compartilhada do Semiárido paraibano, com geração comunitária, captação de água de chuva e forte componente de formação social.",
    77: "Parceria que produz biometano de alta pureza a partir da vinhaça em usinas de Alagoas e Pernambuco, tendo a Usina Cucau como projeto pioneiro.",
    78: "Rede de P&D e capacitação para a transição energética no Nordeste, com TechHubs do SENAI, cooperação Brasil-Alemanha e ecossistema universitário.",
    79: "Programa federal de finanças sustentáveis que mobiliza capital privado internacional via blended finance e hedge cambial; mobilizou cerca de R$ 75 bilhões em três leilões.",
    80: "Primeira emissão de green bonds da Casa dos Ventos (R$ 430 milhões) para financiar a expansão do Complexo Eólico Rio do Vento, no Rio Grande do Norte.",
    81: "Plataforma de filantropia climática que conecta financiadores internacionais a parceiros locais para o desenvolvimento de baixo carbono com justiça social; coorganiza o Fórum de Powershoring.",
    82: "Pioneira em finanças de impacto no Brasil, mobiliza recursos via blended finance e crowdlending; no Nordeste apoia negócios de impacto com crédito acessível.",
}

EIXO_COD = {"Finanças Sustentáveis e Inclusivas": "FSI", "Adensamento Tecnológico": "ADT",
            "Bioeconomia e Sistemas Agroalimentares Adaptados": "BIO", "Transição Energética": "TE",
            "Economia Circular e Solidária": "EC", "Nova Infraestrutura Verde-Azul e Adaptação Climática": "NIVA"}
CRIT = [("relevancia", "Relevância territorial"), ("clima", "Adaptação/mitigação climática"),
        ("cadeias", "Fortalecimento de cadeias produtivas"), ("viabilidade", "Viabilidade operacional"),
        ("replicabilidade", "Replicabilidade/escala"), ("inovacao", "Inovação tecnológica e social"),
        ("investimentos", "Atração de investimentos verdes"), ("inclusao", "Inclusão produtiva"),
        ("equidade", "Equidade de gênero, raça e etnia"), ("governanca", "Governança multinível")]
# Cores dos eixos alinhadas à paleta do Plano Brasil Nordeste (PTE-NE)
EIXO_CORES = [("Finanças Sustentáveis e Inclusivas", "FSI", "#1f4da1"),
              ("Adensamento Tecnológico", "ADT", "#7a3fb8"),
              ("Bioeconomia e Sistemas Agroalimentares Adaptados", "BIO", "#43a047"),
              ("Transição Energética", "TE", "#f37520"),
              ("Economia Circular e Solidária", "EC", "#f6a609"),
              ("Nova Infraestrutura Verde-Azul e Adaptação Climática", "NIVA", "#0d9488")]


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Matriz de avaliação"]

    def val(r, c):
        v = ws.cell(row=r, column=c).value
        if isinstance(v, str):
            v = v.strip()
        return v if v not in ("",) else None

    items = []
    for r in range(4, ws.max_row + 1):
        nm = val(r, 2)
        if not nm or str(nm).strip().upper() == "NOVAS INICIATIVAS":
            continue
        idn = int(val(r, 1))
        if idn in EXCLUDE:
            continue
        nm = NAME_OVERRIDES.get(idn, nm)
        crit = {key: val(r, 22 + k) for k, (key, _) in enumerate(CRIT)}
        eixo = val(r, 16)
        items.append({
            "id": idn, "nome": nm, "objetivo": val(r, 3), "resumo": RESUMO.get(idn), "tematica": val(r, 4), "org": ORG_OVERRIDES.get(idn, val(r, 5)),
            "cnpj": val(r, 6), "fundacao": val(r, 7), "cnae": val(r, 8), "endereco": val(r, 9),
            "lat": val(r, 10), "lon": val(r, 11), "avaliador": val(r, 12),
            "municipio": val(r, 13), "estado": val(r, 14), "biomas": val(r, 15),
            "eixo": eixo, "eixo_cod": EIXO_COD.get(eixo), "eixo_sec": val(r, 17),
            "setor": val(r, 18), "natureza": val(r, 19), "tipo_inst": val(r, 20), "salvaguardas": val(r, 21),
            "criterios": crit, "pontuacao": val(r, 32), "observacoes": val(r, 33),
            "preselecionada": idn in PRE, "pre_ufs": PRE.get(idn, []), "fora_ne": idn in FORA_NE,
        })

    # coordenadas alternativas (rota escolhe a ótima; mapa mostra todas)
    for it in items:
        if it["id"] in ALT_LOCAIS:
            it["locais"] = ALT_LOCAIS[it["id"]]
            it["lat"] = it["locais"][0]["lat"]
            it["lon"] = it["locais"][0]["lon"]

    meta = {"fonte": "PTE2026_matriz_dashboard.xlsx", "total": len(items),
            "preselecionadas": sum(1 for i in items if i["preselecionada"]),
            "eixos": [{"nome": n, "cod": c, "cor": cor} for n, c, cor in EIXO_CORES],
            "criterios": [{"key": k, "label": l} for k, l in CRIT]}
    out = {"meta": meta, "iniciativas": items}
    js = "window.PTE_DATA = " + json.dumps(out, ensure_ascii=False, indent=1) + ";\n"
    with open(OUT, "w", encoding="utf-8") as f:
        f.write(js)
    print(f"OK: {OUT} ({len(js)} bytes) | {meta['total']} iniciativas, {meta['preselecionadas']} pré-selecionadas")


if __name__ == "__main__":
    main()
