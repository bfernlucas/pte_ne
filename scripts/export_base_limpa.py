#!/usr/bin/env python3
"""Exporta uma planilha-base limpa (PTE2026_base_limpa.xlsx) a partir do
arquivo gerado assets/data/iniciativas.js — espelha exatamente o que o
dashboard mostra (iniciativas já limpas, renomeadas e sem duplicatas).
Uso: python3 scripts/export_base_limpa.py
"""
import json, os, openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "assets", "data", "iniciativas.js")
OUT = os.path.join(ROOT, "PTE2026_base_limpa.xlsx")


def load():
    s = open(SRC, encoding="utf-8").read()
    return json.loads(s[s.index("{"):s.rindex("}") + 1])


def main():
    data = load()
    items = data["iniciativas"]
    crit = data["meta"]["criterios"]
    crit_keys = [c["key"] for c in crit]

    base_cols = ["ID", "Iniciativa", "Organização", "Resumo", "Objetivo", "Temática principal",
                 "CNPJ", "Fundação", "CNAE", "Endereço", "Latitude", "Longitude", "Avaliador",
                 "Município", "Estado", "Biomas", "Eixo", "Eixo secundário", "Setor",
                 "Natureza jurídica", "Tipo de organização", "Salvaguardas"]
    headers = base_cols + [c["label"] for c in crit] + ["Nota (0–30)", "Observações",
                 "Pré-selecionada", "UFs (pré-seleção)", "Sede fora do NE"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Iniciativas"
    ws.append(headers)
    for i in items:
        cr = i.get("criterios") or {}
        row = [i.get("id"), i.get("nome"), i.get("org"), i.get("resumo"), i.get("objetivo"),
               i.get("tematica"), i.get("cnpj"), i.get("fundacao"), i.get("cnae"), i.get("endereco"),
               i.get("lat"), i.get("lon"), i.get("avaliador"), i.get("municipio"), i.get("estado"),
               i.get("biomas"), i.get("eixo"), i.get("eixo_sec"), i.get("setor"), i.get("natureza"),
               i.get("tipo_inst"), i.get("salvaguardas")]
        row += [cr.get(k) for k in crit_keys]
        row += [i.get("pontuacao"), i.get("observacoes"),
                "Sim" if i.get("preselecionada") else "Não",
                ", ".join(i.get("pre_ufs") or []),
                "Sim" if i.get("fora_ne") else "Não"]
        ws.append(row)

    # estilo do cabeçalho
    head_fill = PatternFill("solid", fgColor="292D76")
    head_font = Font(color="FFFFFF", bold=True, size=10)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # larguras
    wide = {"Iniciativa": 34, "Organização": 34, "Resumo": 60, "Objetivo": 70, "Temática principal": 22,
            "CNPJ": 20, "Endereço": 30, "Eixo": 30, "Eixo secundário": 26, "Setor": 16,
            "Natureza jurídica": 20, "Tipo de organização": 22, "Salvaguardas": 18,
            "Município": 18, "Biomas": 18, "Observações": 30, "Avaliador": 18}
    for idx, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = wide.get(h, 12)
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            ws.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(OUT)
    print(f"OK: {OUT} | {len(items)} iniciativas, {len(headers)} colunas")


if __name__ == "__main__":
    main()
