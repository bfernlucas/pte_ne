#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Regenerador do dashboard PTE2026.

O que faz:
  1. Lê a planilha-fonte (aba "Dashboard"), preenche campos em branco
     (EXCETO pontuação) e corrige coordenadas malformadas.
  2. Salva a planilha corrigida e um JSON canônico (assets/data/iniciativas.json).
  3. Embute os 52 registros no index.html (substitui `const DATA = [...]`).
  4. (1ª execução) injeta as features de Ranking e Rota de campo no index.html.

Uso:  python3 tools/build_dashboard.py [planilha.xlsx]
"""
import sys, os, json, datetime, re
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = sys.argv[1] if len(sys.argv) > 1 else os.path.join(ROOT, "PTE2026_matriz_dashboard_vf.xlsx")
INDEX = os.path.join(ROOT, "index.html")

# ----------------------------------------------------------------------------
# 1. Preenchimentos (exceto pontuação) e correção de coordenadas
# ----------------------------------------------------------------------------
FILLS = {
    5:  {"municipio": "Rio de Janeiro"},                       # Conexsus (escritório RJ; coords já no RJ)
    8:  {"municipio": "Recife", "estado": "PE", "bioma": "Mata Atlântica"},  # coords já em Recife
    11: {"lat": -15.7942, "lon": -47.8822},                    # SEBRAE — Brasília
    19: {"lat": -15.7942, "lon": -47.8822},                    # Territórios da Cidadania — Brasília
    20: {"lat": -7.2342,  "lon": -39.4094},                    # Programa Redeser — Crato/CE
    27: {"natureza_juridica": "Órgão Público do Poder Executivo Federal"},
    49: {"lat": -23.0032},                                     # corrige -230032 -> -23.0032
}


def s(v):
    return v.replace("\n", " ").strip() if isinstance(v, str) else v


REPO_XLSX = os.path.join(ROOT, "PTE2026_matriz_dashboard_vf.xlsx")


def load_records():
    # data_only=True -> usa os valores calculados (a aba Dashboard puxa de fórmulas).
    wbv = openpyxl.load_workbook(SRC, data_only=True)
    ws = wbv["Dashboard"]
    H = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {h: i + 1 for i, h in enumerate(H)}
    # aplica os preenchimentos direto na planilha e salva uma cópia PORTÁTIL
    # (fórmulas viram valores), garantindo que reexecuções leiam dados estáveis.
    for r in range(2, ws.max_row + 1):
        idv = ws.cell(r, 1).value
        if idv is None:
            continue
        for k, v in FILLS.get(int(idv), {}).items():
            ws.cell(r, col[k]).value = v
    if os.path.abspath(SRC) != os.path.abspath(REPO_XLSX):
        wbv.save(REPO_XLSX)
        print(f"Planilha corrigida (valores) salva em {os.path.basename(REPO_XLSX)}")
    recs = []
    for r in range(2, ws.max_row + 1):
        idv = ws.cell(r, 1).value
        if idv is None:
            continue
        d = {k: ws.cell(r, col[k]).value for k in H}
        df = d.get("data_fundacao")
        if isinstance(df, (datetime.datetime, datetime.date)):
            df = df.strftime("%d/%m/%Y")
        recs.append({
            "id": int(idv),
            "iniciativa": s(d.get("iniciativa")) or "",
            "objetivo": s(d.get("objetivo")) or "",
            "tematica": s(d.get("tematica")) or "",
            "organizacao": s(d.get("organizacao")) or "",
            "cnpj": s(d.get("cnpj")),
            "data_fundacao": df,
            "municipio": s(d.get("municipio")),
            "estado": s(d.get("estado")),
            "bioma": s(d.get("bioma")),
            "eixo": s(d.get("eixo")) or "",
            "setor": s(d.get("setor")),
            "natureza": s(d.get("natureza_juridica")),
            "avaliador": s(d.get("avaliador")),
            "lat": float(d["lat"]) if d.get("lat") not in (None, "") else None,
            "lon": float(d["lon"]) if d.get("lon") not in (None, "") else None,
            "pontuacao": float(d["pontuacao"]) if d.get("pontuacao") not in (None, "") else 0,
        })
    return recs


def main():
    recs = load_records()
    print(f"Registros: {len(recs)}")
    # JSON canônico
    os.makedirs(os.path.join(ROOT, "assets", "data"), exist_ok=True)
    json.dump(recs, open(os.path.join(ROOT, "assets", "data", "iniciativas.json"), "w", encoding="utf-8"),
              ensure_ascii=False, indent=1)
    compact = json.dumps(recs, ensure_ascii=False, separators=(",", ":"))

    html = open(INDEX, encoding="utf-8").read()

    # 3. substitui a linha de dados (DATA é uma única linha física → troca linha inteira)
    lines = html.split("\n")
    repl = 0
    for i, ln in enumerate(lines):
        if ln.startswith("const DATA = ["):
            lines[i] = "const DATA = " + compact + ";"
            repl += 1
    assert repl == 1, f"DATA: esperava 1 linha, achou {repl}"
    html = "\n".join(lines)

    # 4. injeta/atualiza features
    from features import CSS, ROUTEBAR_HTML, PANELS_HTML, FEATURE_JS, PATCHES
    JS_START = "/* PTE-FEATURES-JS:START */"
    JS_END = "/* PTE-FEATURES-JS:END */"
    js_block = JS_START + "\n" + FEATURE_JS + "\n" + JS_END
    if JS_START in html:
        # re-execução: atualiza apenas o bloco JS (CSS/HTML/patches já aplicados)
        html = re.sub(re.escape(JS_START) + ".*?" + re.escape(JS_END),
                      lambda _: js_block, html, count=1, flags=re.S)
        print("FEATURE_JS atualizado.")
    else:
        for old, new in PATCHES:
            cnt = html.count(old)
            assert cnt == 1, f"PATCH não-único ({cnt}): {old[:60]}..."
            html = html.replace(old, new, 1)
        html = html.replace("</style>", CSS + "\n</style>", 1)
        html = html.replace('<div class="kpis" id="kpis"></div>',
                            '<div class="kpis" id="kpis"></div>\n' + ROUTEBAR_HTML, 1)
        html = html.replace('  <div class="table-wrap"><table id="tbl"></table></div>',
                            PANELS_HTML + '\n  <div class="table-wrap"><table id="tbl"></table></div>', 1)
        html = html.replace("buildFilters();render();",
                            js_block + "\nbuildFilters();render();wireRouteControls();", 1)
        print("Features injetadas (1ª vez).")

    open(INDEX, "w", encoding="utf-8").write(html)
    print("index.html atualizado.")


if __name__ == "__main__":
    main()
